import telebot
from telebot import types
import sqlite3
import random
from openpyxl import load_workbook, Workbook

# Токен бота Telegram
TOKEN = '7717241932:AAHF4hFEMPsedHFfK82yp27F5hRvXwG23_0'  # Замени на твой токен
bot = telebot.TeleBot(TOKEN, parse_mode='Markdown')

# Путь к Excel файлу
EXCEL_FILE = "Chetile.xlsx"

# Словари для хранения выбора пользователя
user_column_choice = {}
user_row_choice = {}

# Функция для подключения к базе данных и получения случайного хадиса
def get_random_hadith():
    # Подключаемся к базе данных
    conn = sqlite3.connect('HadithCollection.db')  # Убедись, что путь к базе данных верный
    cursor = conn.cursor()

    # Выбираем случайный хадис
    cursor.execute("SELECT text, narrator, reference FROM Hadiths ORDER BY RANDOM() LIMIT 1")
    hadith = cursor.fetchone()

    conn.close()

    # Формируем текст хадиса для отправки с жирным выделением
    hadith_text = f"`Хадис:` {hadith[0]}\n`Передал(-a):` {hadith[1]}\n`Источник:` {hadith[2]}"
    return hadith_text


@bot.message_handler(commands=['start'])
def send_bot_info(message):
    """Команда /start отправляет информацию о боте и список доступных команд."""
    bot_info = (
        "*Добро пожаловать!*\n"
        "Я бот для управления данными в Chetile.\n\n"
        "*Доступные команды:*\n"
        "`/start` - Информация о боте и список команд.\n"
        "`/get` - Получить случайный хадис и начать редактирование таблицы.\n"
        "`/show` - Отправить Chetile-файл.\n"
        "`/new` - Обнулить все редактируемые ячейки.\n"
        "*Как использовать:*\n"
        "1. Введите `/get`, чтобы получить случайный хадис и выбрать колонку для редактирования.\n"
        "2. Следуйте инструкциям для выбора дня недели и ввода данных.\n"
        "3. Используйте `/show`, чтобы скачать обновленный Chetile-файл.\n"
        "4. Введите `/new`, чтобы обнулить все редактируемые ячейки в таблице."
    )
    bot.send_message(message.chat.id, bot_info)


@bot.message_handler(commands=['get'])
def send_welcome(message):
    """Команда /get отправляет случайный хадис и позволяет начать редактирование."""
    # Получаем случайный хадис
    hadith_text = get_random_hadith()

    # Создаем Inline клавиатуру для выбора столбцов (3x3)
    markup = types.InlineKeyboardMarkup(row_width=3)
    columns = ['KK', 'PR', 'RNK', 'MP3', 'OR', 'TH', 'B/T', 'CV', 'SLVT']
    buttons = [types.InlineKeyboardButton(text=col, callback_data=f"column_{col}") for col in columns]

    # Добавляем кнопки в 3 строки
    markup.add(*buttons)

    # Отправляем случайный хадис с жирным выделением
    bot.send_message(message.chat.id, hadith_text, reply_markup=markup)

@bot.message_handler(commands=['new'])
def reset_excel_cells(message):
    try:
        # Загружаем Excel-файл
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active

        # Диапазоны редактируемых ячеек (например, C4:K10)
        column_range = ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']
        row_range = range(4, 11)  # Ряды с 4 по 10 включительно

        # Обнуляем содержимое указанных ячеек
        for row in row_range:
            for column in column_range:
                ws[f"{column}{row}"] = ""

        # Сохраняем изменения в файл
        wb.save(EXCEL_FILE)

        # Отправляем подтверждение пользователю
        bot.send_message(message.chat.id, "Все редактируемые ячейки были обнулены.")
    except Exception as e:
        bot.send_message(message.chat.id, f"Ошибка при обнулении ячеек: {e}")

# Команда /show для отправки файла Excel
@bot.message_handler(commands=['show'])
def send_excel_file(message):
    # Отправляем пользователю Excel файл
    with open(EXCEL_FILE, 'rb') as file:
        bot.send_document(message.chat.id, file)


# Обработчик для выбора столбцов
@bot.callback_query_handler(func=lambda call: call.data.startswith('column_'))
def handle_column_choice(call):
    column = call.data.split('_')[1]
    user_column_choice[call.message.chat.id] = column

    # Создаем Inline клавиатуру для выбора дня недели
    markup = types.InlineKeyboardMarkup(row_width=3)
    days = ['Понедельник', 'Вторник', 'Среда', 'Четверг', 'Пятница', 'Суббота', 'Воскресенье']
    buttons = [types.InlineKeyboardButton(text=day, callback_data=f"day_{day}") for day in days]

    # Добавляем кнопки в один ряд
    markup.add(*buttons)

    bot.edit_message_text("Выберите день недели:", chat_id=call.message.chat.id, message_id=call.message.message_id,
                          reply_markup=markup)


# Обработчик для выбора дня недели
@bot.callback_query_handler(func=lambda call: call.data.startswith('day_'))
def handle_day_choice(call):
    day = call.data.split('_')[1]
    day_to_row = {
        'Понедельник': 4,
        'Вторник': 5,
        'Среда': 6,
        'Четверг': 7,
        'Пятница': 8,
        'Суббота': 9,
        'Воскресенье': 10
    }

    # Сохраняем выбор дня в словарь пользователя
    user_row_choice[call.message.chat.id] = day_to_row[day]

    # Спрашиваем у пользователя, какое число он хочет записать
    bot.edit_message_text(f"Вы выбрали {day}. Теперь введите число для записи:", chat_id=call.message.chat.id,
                          message_id=call.message.message_id)


# Обработчик для ввода числа
@bot.message_handler(func=lambda message: message.text.isdigit())
def handle_number_input(message):
    column_map = {
        'KK': 'C',
        'PR': 'D',
        'RNK': 'E',
        'MP3': 'F',
        'OR': 'G',
        'TH': 'H',
        'B/T': 'I',
        'CV': 'J',
        'SLVT': 'K'
    }

    # Получаем выбранный столбец и строку для пользователя
    if message.chat.id in user_column_choice and message.chat.id in user_row_choice:
        column_letter = column_map[user_column_choice[message.chat.id]]
        row_number = user_row_choice[message.chat.id]
        value = message.text

        # Формируем адрес ячейки в формате, например, C4
        cell = f'{column_letter}{row_number}'

        # Открываем Excel файл и записываем данные
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        ws[cell] = value
        wb.save(EXCEL_FILE)

        # Отправляем подтверждение пользователю
        bot.send_message(message.chat.id, f"Значение {value} записано в ячейку {cell}.")
    else:
        bot.send_message(message.chat.id, "Пожалуйста, сначала выберите столбец и день недели.")


# Запуск бота
if __name__ == "__main__":
    # Проверяем, существует ли файл Excel. Если нет, создаем новый.
    try:
        open(EXCEL_FILE, 'r').close()
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        wb.save(EXCEL_FILE)

    bot.polling(none_stop=True)
